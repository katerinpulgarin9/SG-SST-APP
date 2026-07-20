# -*- coding: utf-8 -*-
"""Carga el catalogo desde FG-FOR-SST-01 CONTROL MAESTRO DOCUMENTAL.

Recodifica cada fila al esquema inicial del SG-SST:
  SST-SG-[ABR]-[NNN]
donde ABR se infiere del nombre/familia y NNN es el consecutivo unico del maestro.
"""
from __future__ import annotations

import re
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parent.parent
CONTROL_MAESTRO = ROOT / "data" / "FG-FOR-SST-01_CONTROL_MAESTRO_DOCUMENTAL.xlsx"
CONTROL_MAESTRO_ALT = ROOT.parent / "FG-FOR-SST-01 - CONTROL MAESTRO DOCUMENTAL.xlsx"

CODIGO_SISTEMA = "SST-SG"


def _norm(texto: str) -> str:
    """Mayusculas sin acentos para clasificar nombres del maestro."""
    t = (texto or "").upper()
    for a, b in (
        ("\u00c1", "A"),
        ("\u00c9", "E"),
        ("\u00cd", "I"),
        ("\u00d3", "O"),
        ("\u00da", "U"),
        ("\u00d1", "N"),
    ):
        t = t.replace(a, b)
    # combinar acentos (NFD residual)
    for ch in ("\u0301", "\u0300", "\u0303", "\u0308"):
        t = t.replace(ch, "")
    return t


def _infer_familia(nombre: str, formato: str) -> str:
    n = _norm(nombre)
    if "FORMATO" in n or "ENCUESTA" in n or "INSPECCION" in n or "PREOPERACIONAL" in n:
        return "formato"
    if "ACTA" in n:
        return "acta"
    if "PROCEDIMIENTO" in n or "PROCED." in n:
        return "procedimiento"
    if "PROGRAMA" in n:
        return "programa"
    if "PLAN " in n or n.startswith("PLAN") or "PLANES" in n:
        return "plan"
    if "MATRIZ" in n or "IPER" in n or "IPVR" in n:
        return "matriz"
    if "EVALUACION" in n or "AUTOEVALU" in n or "AUDITOR" in n:
        return "evaluacion"
    if "MANUAL" in n or "ESTANDAR" in n:
        return "procedimiento"
    if n.startswith("POL ") or n == "POL POLITICA" or (
        "POLITICA" in n and "FORMATO" not in n and "CONSENTIMIENTO" not in n
    ):
        return "politica"
    if "POLITICA" in n:
        return "politica"
    if "PRESUPUESTO" in n or "INDICADOR" in n or "ESTADISTIC" in n:
        return "matriz" if formato == "Excel" else "formato"
    return "formato" if formato == "Excel" else "procedimiento"


def _infer_abreviatura(nombre: str, familia: str) -> str:
    mapa = {
        "politica": "POL",
        "acta": "ACT",
        "procedimiento": "PRO",
        "programa": "PRG",
        "plan": "PL",
        "matriz": "MT",
        "evaluacion": "E",
        "formato": "FR",
    }
    return mapa.get(familia, "FR")


def _parse_numero(codigo: str) -> int:
    m = re.search(r"(\d+)$", codigo or "")
    return int(m.group(1)) if m else 0


def codigo_sst(abreviatura: str, numero: int) -> str:
    """Construye SST-SG-[ABR]-[NNN] con NNN de 3 digitos."""
    abr = re.sub(r"[^A-Za-z0-9]", "", (abreviatura or "FR").upper())[:12] or "FR"
    return f"{CODIGO_SISTEMA}-{abr}-{numero:03d}"


# Herramientas / equipos: formatos preoperacionales y carpeta dedicada
_HERRAMIENTA_NOMBRE = (
    "MOTOSIERRA",
    "GUADAN",
    "MACHETE",
    "HACHA",
    "TIJERA",
    "ASPERSOR",
    "DESBROZ",
    "COMPRESOR",
    "PLANTA ELECTRICA",
    "HERRAMIENTAS DE CORTE",
    "HERRAMIENTAS MANUALES",
    "PREOPERACIONAL",
    "MAQUINAS, EQUIPOS Y HERRAMIENT",
)
_HERRAMIENTA_RUTA = ("_NUEVOS_FORMATOS_HERRAMIENTAS", "HERRAMIENTAS/")


def _nombre_clave(nombre: str) -> str:
    n = _norm(nombre)
    return " ".join("".join(ch if ch.isalnum() or ch.isspace() else " " for ch in n).split())


def es_documento_herramienta(nombre: str, referencia: str = "") -> bool:
    """True si el documento es formato/estandar de herramienta o equipo especifico."""
    ref = _norm(referencia).replace("\\", "/")
    if any(k in ref for k in _HERRAMIENTA_RUTA):
        return True
    n = _norm(nombre)
    # Nombre casi solo el equipo (ej. ASPERSORA, DESBROZADORA)
    if n in {"ASPERSOR", "ASPERSORA", "DESBROZADORA", "DESBROZADOR", "GUADANA", "MOTOSIERRA"}:
        return True
    if "PREOPERACIONAL" in n:
        return True
    if "HERRAMIENT" in n and any(
        k in n for k in ("MOTOSIERRA", "GUADAN", "CORTE", "DESBROCE", "MANUAL", "MAQUINA")
    ):
        return True
    if any(k in n for k in ("MOTOSIERRA", "GUADAN", "COMPRESOR", "PLANTA ELECTRICA")) and (
        "INSPECCION" in n or "ESTANDAR" in n or "FORMATO" in n or "MANEJO DE" in n
    ):
        return True
    return False


def filtrar_catalogo(docs: list[dict[str, Any]]) -> list[dict[str, Any]]:
    """Quita documentos de herramientas y duplicados (deja la primera aparicion)."""
    vistos: set[str] = set()
    limpios: list[dict[str, Any]] = []
    for d in sorted(docs, key=lambda x: int(x.get("numero") or 0)):
        if es_documento_herramienta(d.get("nombre", ""), d.get("referencia", "")):
            continue
        clave = _nombre_clave(d.get("nombre", ""))
        if not clave or clave in vistos:
            continue
        vistos.add(clave)
        limpios.append(d)
    return limpios


def cargar_control_maestro(
    path: Path | None = None,
    *,
    filtrar: bool = True,
) -> list[dict[str, Any]]:
    fuente = path or CONTROL_MAESTRO
    if not fuente.exists():
        fuente = CONTROL_MAESTRO_ALT
    if not fuente.exists():
        raise FileNotFoundError(
            "No se encontro FG-FOR-SST-01 CONTROL MAESTRO DOCUMENTAL.xlsx"
        )

    wb = load_workbook(fuente, data_only=True)
    ws = wb.active
    docs: list[dict[str, Any]] = []
    for r in range(2, ws.max_row + 1):
        codigo_origen = ws.cell(r, 1).value
        nombre = ws.cell(r, 2).value
        tipo = ws.cell(r, 3).value
        ruta = ws.cell(r, 4).value
        estado = ws.cell(r, 5).value
        if not codigo_origen or not nombre:
            continue
        origen_s = str(codigo_origen).strip()
        nombre_s = str(nombre).strip()
        if nombre_s in {"~1", "~$", "-"} or len(nombre_s) < 2:
            nombre_s = f"Documento {origen_s}"
        ext = str(tipo or "").strip().upper()
        formato = "Excel" if "XLS" in ext else "Word"
        familia = _infer_familia(nombre_s, formato)
        abr = _infer_abreviatura(nombre_s, familia)
        numero = _parse_numero(origen_s)
        if numero <= 0:
            continue
        docs.append(
            {
                "codigo": codigo_sst(abr, numero),
                "codigo_origen": origen_s,
                "abreviatura": abr,
                "nombre": nombre_s,
                "formato": formato,
                "familia": familia,
                "referencia": str(ruta or "").strip(),
                "estado": str(estado or "").strip() or "CONFORME",
                "numero": numero,
            }
        )
    if filtrar:
        docs = filtrar_catalogo(docs)
    return docs
